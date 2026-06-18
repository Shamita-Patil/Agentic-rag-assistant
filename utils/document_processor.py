import os

from pypdf import PdfReader

from docx import Document

import openpyxl


def load_pdf(file_path):

    docs = []

    reader = PdfReader(file_path)

    for page_num, page in enumerate(
        reader.pages
    ):

        text = page.extract_text()

        if text:

            docs.append(
                {
                    "content": text,
                    "metadata": {
                        "source":
                        os.path.basename(
                            file_path
                        ),
                        "page":
                        page_num + 1
                    }
                }
            )

    return docs


def load_txt(file_path):

    with open(
        file_path,
        "r",
        encoding="utf-8"
    ) as f:

        text = f.read()

    return [
        {
            "content": text,
            "metadata": {
                "source":
                os.path.basename(
                    file_path
                )
            }
        }
    ]


def load_docx(file_path):

    doc = Document(
        file_path
    )

    text = "\n".join(
        [
            p.text
            for p in doc.paragraphs
        ]
    )

    return [
        {
            "content": text,
            "metadata": {
                "source":
                os.path.basename(
                    file_path
                )
            }
        }
    ]


def load_xlsx(file_path):

    workbook = (
        openpyxl.load_workbook(
            file_path
        )
    )

    rows = []

    for sheet in workbook.sheetnames:

        ws = workbook[sheet]

        for row in ws.iter_rows(
            values_only=True
        ):

            values = [
                str(v)
                for v in row
                if v is not None
            ]

            if values:

                rows.append(
                    " ".join(
                        values
                    )
                )

    return [
        {
            "content":
            "\n".join(rows),
            "metadata": {
                "source":
                os.path.basename(
                    file_path
                )
            }
        }
    ]


def load_documents(
    document_folder
):

    documents = []

    for filename in os.listdir(
        document_folder
    ):

        path = os.path.join(
            document_folder,
            filename
        )

        if filename.lower().endswith(
            ".pdf"
        ):

            documents.extend(
                load_pdf(path)
            )

        elif filename.lower().endswith(
            ".txt"
        ):

            documents.extend(
                load_txt(path)
            )

        elif filename.lower().endswith(
            ".docx"
        ):

            documents.extend(
                load_docx(path)
            )

        elif filename.lower().endswith(
            ".xlsx"
        ):

            documents.extend(
                load_xlsx(path)
            )

    return documents